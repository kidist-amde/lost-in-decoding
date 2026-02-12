# tools/query_attack_results_pag_vs_dense.py

import json
import os
from typing import Any, Dict, List, Tuple

import numpy as np
import pandas as pd
import matplotlib.pyplot as plt
import matplotlib.colors as mcolors
from openpyxl import load_workbook

# Set font to serif (Times New Roman style)

# Set font to serif (Times New Roman style)
plt.rcParams['font.family'] = 'serif'
plt.rcParams['font.serif'] = ['Times New Roman', 'DejaVu Serif', 'STIXGeneral']
plt.rcParams['font.size'] = 22
plt.rcParams['axes.titlesize'] = 22
plt.rcParams['axes.labelsize'] = 24
plt.rcParams['xtick.labelsize'] = 22
plt.rcParams['ytick.labelsize'] = 22
plt.rcParams['legend.fontsize'] = 22



def safe_read_attack_xlsx(path: str) -> Tuple[List[str], List[List[Any]]]:
    """Read xlsx as raw table data to avoid pandas parser issues."""
    wb = load_workbook(path, data_only=True, read_only=True)
    ws = wb[wb.sheetnames[0]]

    rows = [list(r) for r in ws.iter_rows(values_only=True)]
    rows = [r for r in rows if any(v is not None for v in r)]
    if not rows:
        return [], []

    n_cols = max(len(r) for r in rows)

    def normalize_row(row: List[Any]) -> List[Any]:
        row = list(row[:n_cols])
        if len(row) < n_cols:
            row.extend([None] * (n_cols - len(row)))
        return row

    rows = [normalize_row(r) for r in rows]
    header = [str(v).strip() if v is not None else f"col_{i}" for i, v in enumerate(rows[0])]
    return header, rows[1:]


def main() -> None:
    model_list = [
        "qwen3",
        "embeddinggemma",
        "tas_b",
        "nomic_v2",
    ]

    model_display_names = {
        "qwen3": "Qwen3",
        "embeddinggemma": "EmbeddingGemma",
        "tas_b": "TAS-B",
        "nomic_v2": "Nomic-v2",
    }

    dataset_list = ["msmarco"]
    split_list = ["dev", "trec_dl19", "trec_dl20"]

    split_display_names = {
        "dev": "MS MARCO (Dev)",
        "trec_dl19": "TREC DL 2019",
        "trec_dl20": "TREC DL 2020",
    }

    attack_method_list = [
        "none",
        "mispelling",
        "ordering",
        "synonym",
        "paraphrase",
        "naturality",
    ]

    attack_method_display_names = {
        "none": "Clean",
        "mispelling": "Misspelling",
        "ordering": "Reordering",
        "synonym": "Synonymizing",
        "paraphrase": "Paraphrasing",
        "naturality": "Naturalizing",
    }

    attack_methods_for_eval = [m for m in attack_method_list if m != "none"]
    seed_list = [1999, 5, 27, 2016, 2026]

    scores_base_path = "<PATH_TO_ATTACK_SCORES>"
    os.makedirs("tools/excel_results", exist_ok=True)

    all_data: Dict[str, Dict[str, Dict[str, Any]]] = {}
    all_metrics: Dict[str, str] = {}

    for dataset in dataset_list:
        for split in split_list:
            target_metric = "MRR@10" if split == "dev" else "NDCG@10"

            key = f"{dataset}_{split}"
            all_metrics[key] = target_metric
            data: Dict[str, Dict[str, Any]] = {}

            for model in model_list:
                data[model] = {}

                clean_metric_list: List[float] = []
                for seed in seed_list:
                    path = (
                        f"{scores_base_path}/{model}/{dataset}_{split}/"
                        f"attack_method_none_seed_{seed}_attacked_num_50/metrics_scores_and_asr.json"
                    )
                    try:
                        with open(path, "r") as f:
                            scores = json.load(f)
                        clean_metric_list.append(float(scores[target_metric]))
                    except FileNotFoundError:
                        print(f"[WARN] File not found: {path}")

                if not clean_metric_list:
                    print(f"[WARN] No clean data for model={model}, split={split}")
                    continue

                clean_mean = float(np.mean(clean_metric_list))
                clean_std = float(np.std(clean_metric_list))
                data[model]["clean_mean"] = clean_mean
                data[model]["clean_std"] = clean_std
                print(f"model={model}, split={split}, clean {target_metric}: {clean_mean:.4f} ± {clean_std:.4f}")

                for attack_method in attack_methods_for_eval:
                    decrease_rate_list: List[float] = []
                    for seed_idx, seed in enumerate(seed_list):
                        path = (
                            f"{scores_base_path}/{model}/{dataset}_{split}/"
                            f"attack_method_{attack_method}_seed_{seed}_attacked_num_50/metrics_scores_and_asr.json"
                        )
                        try:
                            with open(path, "r") as f:
                                scores = json.load(f)
                            attacked_val = float(scores[target_metric])
                            if seed_idx < len(clean_metric_list):
                                clean_val = clean_metric_list[seed_idx]
                                dr = (clean_val - attacked_val) / clean_val
                                decrease_rate_list.append(float(dr))
                        except FileNotFoundError:
                            print(f"[WARN] File not found: {path}")

                    if not decrease_rate_list:
                        print(f"[WARN] No attacked data for model={model}, split={split}, attack={attack_method}")
                        continue

                    dr_mean = float(np.mean(decrease_rate_list))
                    dr_std = float(np.std(decrease_rate_list))
                    data[model][attack_method] = (dr_mean, dr_std)
                    print(f"  attack={attack_method}, decrease_rate: {dr_mean:.4f} ± {dr_std:.4f}")

            all_data[key] = data

    for dataset in dataset_list:
        for split in split_list:
            key = f"{dataset}_{split}"
            data = all_data[key]

            columns: List[str] = []
            for model in model_list:
                display = model_display_names[model]
                columns.append(display)
                columns.append(f"{display}-std")

            rows: List[List[float]] = []
            row_labels: List[str] = []

            metric = all_metrics[key]
            row: List[float] = []
            for model in model_list:
                if "clean_mean" in data[model]:
                    row.append(round(float(data[model]["clean_mean"]), 4))
                    row.append(round(float(data[model]["clean_std"]), 4))
                else:
                    row.append(np.nan)
                    row.append(np.nan)
            rows.append(row)
            row_labels.append(f"Clean ({metric})")

            for attack_method in attack_methods_for_eval:
                row = []
                for model in model_list:
                    if attack_method in data[model]:
                        row.append(round(float(data[model][attack_method][0]), 4))
                        row.append(round(float(data[model][attack_method][1]), 4))
                    else:
                        row.append(np.nan)
                        row.append(np.nan)
                rows.append(row)
                row_labels.append(attack_method_display_names[attack_method])

            df = pd.DataFrame(rows, columns=columns, index=row_labels)
            df.index.name = "Attack Method"

            excel_path = f"tools/excel_results/pag_{dataset}_{split}_query_attack.xlsx"
            df.to_excel(excel_path, index_label="Attack Method")
            print(f"\nSaved Excel: {excel_path}")

    summary_rows: List[Dict[str, str]] = []
    summary_index: List[Tuple[str, str]] = []

    for dataset in dataset_list:
        for split in split_list:
            key = f"{dataset}_{split}"
            data = all_data[key]
            split_display = split_display_names[split]

            row: Dict[str, str] = {}
            for model in model_list:
                display = model_display_names[model]
                if "clean_mean" in data[model]:
                    row[display] = f"{data[model]['clean_mean']:.4f} ± {data[model]['clean_std']:.4f}"
                else:
                    row[display] = "N/A"
            summary_rows.append(row)
            summary_index.append((split_display, f"Clean ({all_metrics[key]})"))

            for attack_method in attack_methods_for_eval:
                row = {}
                for model in model_list:
                    display = model_display_names[model]
                    if attack_method in data[model]:
                        mean_val = data[model][attack_method][0]
                        std_val = data[model][attack_method][1]
                        row[display] = f"{mean_val:.4f} ± {std_val:.4f}"
                    else:
                        row[display] = "N/A"
                summary_rows.append(row)
                summary_index.append((split_display, attack_method_display_names[attack_method]))

    multi_index = pd.MultiIndex.from_tuples(summary_index, names=["Split", "Attack Method"])
    summary_df = pd.DataFrame(summary_rows, index=multi_index)

    summary_excel_path = "tools/excel_results/pag_msmarco_all_splits_query_attack_summary.xlsx"
    summary_df.to_excel(summary_excel_path)
    print(f"\nSaved summary Excel: {summary_excel_path}")

    print("\n" + "=" * 100)
    print("Query Attack Results Summary for PAG (MS MARCO)")
    print("=" * 100)

    for dataset in dataset_list:
        for split in split_list:
            key = f"{dataset}_{split}"
            data = all_data[key]
            split_display = split_display_names[split]

            print(f"\n--- {split_display} ---")
            header = f"{'Attack Method':<20}"
            for model in model_list:
                header += f"{model_display_names[model]:>25}"
            print(header)
            print("-" * (20 + 25 * len(model_list)))

            metric = all_metrics[key]
            line = f"{'Clean (' + metric + ')':<20}"
            for model in model_list:
                if "clean_mean" in data[model]:
                    line += f"{data[model]['clean_mean']:>18.4f} ± {data[model]['clean_std']:<6.4f}"
                else:
                    line += f"{'N/A':>25}"
            print(line)

            for attack_method in attack_methods_for_eval:
                line = f"{attack_method_display_names[attack_method]:<20}"
                for model in model_list:
                    if attack_method in data[model]:
                        mean_val = data[model][attack_method][0]
                        std_val = data[model][attack_method][1]
                        line += f"{mean_val:>18.4f} ± {std_val:<6.4f}"
                    else:
                        line += f"{'N/A':>25}"
                print(line)

    print("\n" + "=" * 100)
    print("Note: Clean row shows absolute metric value, attack rows show Drop Rate (ratio)")
    print("      Dev uses MRR@10, TREC DL 2019/2020 use NDCG@10")
    print("=" * 100)


def draw_query_attack_results_pag():
    pag_data = {
        "trec_dl19": {
            "clean": (0.669, 0.000, None, None),
            "mispelling": (0.452, 0.012, 0.217, 0.012),
            "ordering": (0.654, 0.009, 0.014, 0.009),
            "synonym": (0.526, 0.018, 0.143, 0.018),
            "paraphrase": (0.596, 0.012, 0.073, 0.012),
            "naturality": (0.626, 0.000, 0.043, 0.000),
        },
        "trec_dl20": {
            "clean": (0.621, 0.000, None, None),
            "mispelling": (0.461, 0.019, 0.161, 0.019),
            "ordering": (0.607, 0.009, 0.014, 0.009),
            "synonym": (0.508, 0.007, 0.114, 0.007),
            "paraphrase": (0.512, 0.019, 0.109, 0.019),
            "naturality": (0.598, 0.000, 0.023, 0.000),
        },
        "dev": {
            "clean": (0.362, 0.000, None, None),
            "mispelling": (0.215, 0.003, 0.147, 0.003),
            "ordering": (0.350, 0.001, 0.012, 0.001),
            "synonym": (0.268, 0.002, 0.094, 0.002),
            "paraphrase": (0.300, 0.002, 0.062, 0.002),
            "naturality": (0.342, 0.000, 0.020, 0.000),
        },
    }

    split_metric = {"trec_dl19": "NDCG@10", "trec_dl20": "NDCG@10", "dev": "MRR@10"}
    split_display = {"trec_dl19": "TREC DL 2019", "trec_dl20": "TREC DL 2020", "dev": "MS MARCO (Dev)"}
    attack_display = {
        "clean": "Clean",
        "mispelling": "Misspelling",
        "ordering": "Reordering",
        "synonym": "Synonymizing",
        "paraphrase": "Paraphrasing",
        "naturality": "Naturalizing",
    }

    print("=" * 70)
    print("PAG (Stage 2: End-to-end) Data from Table 4")
    print("=" * 70)
    for split in ["trec_dl19", "trec_dl20", "dev"]:
        metric = split_metric[split]
        print(f"\n--- {split_display[split]} ({metric}) ---")
        print(f"{'Variation':<15} {'Value':>20} {'Delta':>20}")
        print("-" * 55)
        for variation in ["clean", "mispelling", "ordering", "synonym", "paraphrase", "naturality"]:
            mean, std, d_mean, d_std = pag_data[split][variation]
            val_str = f"{mean:.3f} ± {std:.3f}"
            delta_str = f"{d_mean:.3f} ± {d_std:.3f}" if d_mean is not None else "–"
            print(f"{attack_display[variation]:<15} {val_str:>20} {delta_str:>20}")

    pag_decrease_rate = {
        "trec_dl19": {
            "clean_mean": 0.6690,
            "clean_std": 0.0000,
            "mispelling": (0.3243, 0.0179),
            "ordering": (0.0209, 0.0134),
            "synonym": (0.2138, 0.0269),
            "paraphrase": (0.1091, 0.0179),
            "naturality": (0.0643, 0.0000),
        },
        "trec_dl20": {
            "clean_mean": 0.6210,
            "clean_std": 0.0000,
            "mispelling": (0.2593, 0.0306),
            "ordering": (0.0225, 0.0145),
            "synonym": (0.1836, 0.0113),
            "paraphrase": (0.1755, 0.0306),
            "naturality": (0.0370, 0.0000),
        },
        "dev": {
            "clean_mean": 0.3620,
            "clean_std": 0.0000,
            "mispelling": (0.4061, 0.0083),
            "ordering": (0.0331, 0.0028),
            "synonym": (0.2597, 0.0055),
            "paraphrase": (0.1713, 0.0055),
            "naturality": (0.0552, 0.0000),
        },
    }

    print("\n" + "=" * 70)
    print("PAG Decrease Rate (consistent with main(): dr = delta / clean)")
    print("=" * 70)
    for split in ["trec_dl19", "trec_dl20", "dev"]:
        metric = split_metric[split]
        print(f"\n--- {split_display[split]} ({metric}) ---")
        print(f"  Clean: {pag_decrease_rate[split]['clean_mean']:.4f} ± {pag_decrease_rate[split]['clean_std']:.4f}")
        for variation in ["mispelling", "ordering", "synonym", "paraphrase", "naturality"]:
            dr_mean, dr_std = pag_decrease_rate[split][variation]
            print(f"  {attack_display[variation]:<15} decrease_rate: {dr_mean:.4f} ± {dr_std:.4f}")

    ripor_decrease_rate = {
        "trec_dl19": {
            "clean_mean": 0.5604,
            "clean_std": 0.0000,
            "mispelling": (0.3231, 0.0492),
            "ordering": (0.0260, 0.0203),
            "synonym": (0.2311, 0.1171),
            "paraphrase": (0.1147, 0.0266),
            "naturality": (0.0480, 0.0000),
        },
        "trec_dl20": {
            "clean_mean": 0.5587,
            "clean_std": 0.0000,
            "mispelling": (0.3877, 0.0208),
            "ordering": (0.0260, 0.0122),
            "synonym": (0.1997, 0.0349),
            "paraphrase": (0.1738, 0.0415),
            "naturality": (0.0447, 0.0000),
        },
        "dev": {
            "clean_mean": 0.2820,
            "clean_std": 0.0000,
            "mispelling": (0.4823, 0.0028),
            "ordering": (0.0436, 0.0039),
            "synonym": (0.2957, 0.0046),
            "paraphrase": (0.1897, 0.0064),
            "naturality": (0.0465, 0.0000),
        },
    }

    print("\n" + "=" * 70)
    print("RIPOR Decrease Rate (dr = delta / clean)")
    print("=" * 70)
    for split in ["trec_dl19", "trec_dl20", "dev"]:
        metric = split_metric[split]
        print(f"\n--- {split_display[split]} ({metric}) ---")
        print(f"  Clean: {ripor_decrease_rate[split]['clean_mean']:.4f} ± {ripor_decrease_rate[split]['clean_std']:.4f}")
        for variation in ["mispelling", "ordering", "synonym", "paraphrase", "naturality"]:
            dr_mean, dr_std = ripor_decrease_rate[split][variation]
            print(f"  {attack_display[variation]:<15} decrease_rate: {dr_mean:.4f} ± {dr_std:.4f}")

    # ===== Load baseline model data from Excel (ROBUST openpyxl reader; no pd.read_excel) =====
    baseline_model_list = ["tas_b", "embeddinggemma", "nomic_v2", "qwen3"]
    baseline_display_names = {
        "qwen3": "Qwen3-8B",
        "embeddinggemma": "EmbeddingGemma",
        "tas_b": "TAS-B",
        "nomic_v2": "Nomic-v2",
    }

    model_column_aliases = {
        "qwen3": ["Qwen3-8B", "Qwen3"],
        "embeddinggemma": ["EmbeddingGemma"],
        "tas_b": ["TAS-B"],
        "nomic_v2": ["Nomic-v2"],
    }

    attack_methods = ["mispelling", "ordering", "synonym", "paraphrase", "naturality"]
    attack_display_to_key = {
        "Misspelling": "mispelling",
        "Reordering": "ordering",
        "Synonymizing": "synonym",
        "Paraphrasing": "paraphrase",
        "Naturalizing": "naturality",
    }

    baseline_data: Dict[str, Dict[str, Dict[str, Tuple[float, float]]]] = {
        split: {m: {} for m in baseline_model_list} for split in ["trec_dl19", "trec_dl20", "dev"]
    }

    for split in ["trec_dl19", "trec_dl20", "dev"]:
        excel_path = f"tools/excel_results/pag_msmarco_{split}_query_attack.xlsx"
        if not os.path.exists(excel_path):
            print(f"[WARN] Excel not found: {excel_path}, run main() first.")
            continue

        header, data_rows = safe_read_attack_xlsx(excel_path)
        print(f"Loaded baseline data from: {excel_path}")
        if not header:
            continue

        col_idx = {name: i for i, name in enumerate(header)}
        attack_col_idx = col_idx.get("Attack Method", 0)

        for row in data_rows:
            row_label = str(row[attack_col_idx]).strip() if row[attack_col_idx] is not None else ""
            attack_key = attack_display_to_key.get(row_label)
            if attack_key is None:
                continue  # skip Clean row

            for model_key in baseline_model_list:
                mean_col = None
                std_col = None
                for alias in model_column_aliases[model_key]:
                    alias_std = f"{alias}-std"
                    if alias in col_idx and alias_std in col_idx:
                        mean_col = alias
                        std_col = alias_std
                        break
                if mean_col is None or std_col is None:
                    continue

                mean_val = row[col_idx[mean_col]]
                std_val = row[col_idx[std_col]]
                if mean_val is None:
                    continue

                try:
                    mean_f = float(mean_val)
                    std_f = float(std_val) if std_val is not None else 0.0
                except (TypeError, ValueError):
                    continue

                baseline_data[split][model_key][attack_key] = (mean_f, std_f)


    # ========== Plot: 1x3 subplots ==========
    # All "models": 4 baselines + RIPOR + PAG
    all_model_keys = baseline_model_list + ["RIPOR", "PAG"]
    all_display_names = {**baseline_display_names, "RIPOR": "RIPOR", "PAG": "PAG"}
    n_models = len(all_model_keys)

    split_order = ["trec_dl19", "trec_dl20", "dev"]
    subplot_labels = ["(a)", "(b)", "(c)"]

    fig, axes = plt.subplots(1, 3, figsize=(34, 8), sharey=False)

    x = np.arange(len(attack_methods))
    width = 0.13

    def lighten_color(color: Any, amount: float = 0.35) -> Tuple[float, float, float]:
        rgb = np.array(mcolors.to_rgb(color))
        return tuple(rgb + (1.0 - rgb) * amount)

    patterns = ['', '/', '\\', 'x', '+', 'o']
    base_colors = [lighten_color(c) for c in plt.cm.Pastel1.colors[:n_models]]
    raw_model_colors = {
        "tas_b": "#87cefa",          # light sky blue
        "embeddinggemma": "#98ffcc", # mint
        "nomic_v2": "#fa8072",       # salmon
        "qwen3": "#ffd966",          # yellow
        "RIPOR": "#c8a2c8",          # purple
    }
    model_colors = {k: lighten_color(v) for k, v in raw_model_colors.items()}

    for idx, split in enumerate(split_order):
        ax = axes[idx]
        metric = split_metric[split]
        max_with_error = 0.0

        for i, model_key in enumerate(all_model_keys):
            means = []
            stds = []
            for attack in attack_methods:
                if model_key == "PAG":
                    src = pag_decrease_rate[split]
                elif model_key == "RIPOR":
                    src = ripor_decrease_rate[split]
                else:
                    src = baseline_data[split][model_key]

                if attack in src:
                    dr_mean, dr_std = src[attack]
                    means.append(dr_mean * 100)
                    stds.append(dr_std * 100)
                else:
                    means.append(0)
                    stds.append(0)

            if means:
                max_with_error = max(
                    max_with_error,
                    max(m + s for m, s in zip(means, stds)),
                )

            offset = (i - n_models / 2 + 0.5) * width
            bar_color = model_colors.get(model_key, base_colors[i])
            ax.bar(x + offset, means, width, yerr=stds,
                   label=all_display_names[model_key] if idx == 0 else "",
                   capsize=3, color=bar_color,
                   hatch=patterns[i % len(patterns)],
                   edgecolor='dimgray', linewidth=0.4,
                   error_kw={'elinewidth': 1, 'capthick': 1})

        ax.set_xlabel(
            f'{subplot_labels[idx]} {metric} Drop Rate on {split_display[split]}',
            fontsize=20,
            labelpad=10,
        )
        ax.set_xticks(x)
        ax.set_xticklabels(
            [attack_display[m] for m in attack_methods],
            rotation=0,
            ha="center",
            fontsize=18,
        )
        ax.axhline(y=0, color='gray', linestyle='--', linewidth=0.5)
        ax.yaxis.grid(True, linestyle='--', alpha=0.5, color='gray', linewidth=0.5)
        ax.set_axisbelow(True)
        ax.set_ylabel('Drop Rate (%)', fontsize=20, labelpad=4)
        ax.margins(x=0.06, y=0.03)
        base_tops = [35, 40, 50]
        upper = max(base_tops[idx], max_with_error + 2.0)
        ax.set_ylim(0, upper)

    # Shared legend
    handles, labels = axes[0].get_legend_handles_labels()
    fig.legend(handles, labels, loc='upper center', bbox_to_anchor=(0.5, 1.02),
               ncol=n_models, frameon=False, columnspacing=1.0,
               fontsize=22, handlelength=2, handleheight=1.5)

    plt.tight_layout()
    plt.subplots_adjust(top=0.80, bottom=0.24, wspace=0.20)
    os.makedirs("figures", exist_ok=True)
    plt.savefig("figures/pag_msmarco_3splits_decrease_rate.pdf", bbox_inches="tight", pad_inches=0.02)
    plt.savefig("figures/pag_msmarco_3splits_decrease_rate.png", dpi=600, bbox_inches="tight", pad_inches=0.02)
    plt.close()
    print("\nSaved: figures/pag_msmarco_3splits_decrease_rate.pdf & .png")

    return pag_data, pag_decrease_rate


if __name__ == "__main__":
    # main()  # stage 1 baseline results
    draw_query_attack_results_pag()  # stage 2 PAG results
